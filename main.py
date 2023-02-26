

from tkinter import *
from tkinter import ttk
from tkinter import font
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import webbrowser
import matplotlib.pyplot as plt
from matplotlib.ticker import LinearLocator
from datetime import date
import sqlite3
import shutil
import win10toast
import pylab
from parser import *
import threading
import time
#открытие db
conn=sqlite3.connect(r'discounts.db', check_same_thread=False)
cur=conn.cursor()
#уведомление
toast = win10toast.ToastNotifier()
#завершение работы программы
def window_closed():
    window.destroy()
    os.abort()
#запуск парсера всех разделов
def start_all():
    start_categories(0)
    start_products(0)
    if var1.get()==1:
        message='Произошло общее обновление данных разделов "категории" и "товары".'
        toast.show_toast(title='Обновление данных',msg=message,threaded=False,duration=5,icon_path='invisible.ico')
#запуск парсера раздела "категории"
def start_categories(notify):
    refresh_Button_categories.config(text='Загрузка...',state='disabled')
    refresh_Button_products.config(text='Загрузка...',state='disabled')
    renew_all_button.config(text='Загрузка...',state='disabled')
    prev_disc=len(cur.execute("SELECT * FROM categories;").fetchall())
    result = cur.execute("SELECT * FROM target_categories;")
    names_categories = list(element[0] for element in result.fetchall())
    categories_values = []
    for category in names_categories:
        for element in get_discounts_mvideo(category,0): categories_values.append(element)
        for element in get_discounts_dns(category, 0): categories_values.append(element)
        #for element in get_discounts_citilink(category): categories_values.append(element)
    current_date = date.today().strftime('%d.%m.%Y')
    cur.execute("DELETE from categories")
    for element in categories_values:
        element = element + [current_date]
        cur.execute("INSERT INTO categories VALUES (?,?,?,?,?,?,?,?)", element)
    conn.commit()
    cur_disc = len(cur.execute("SELECT * FROM categories;").fetchall())
    fill_categories_variables()
    try:
        id_table = table_categories.get_children()[0]
        value=table_categories.item(id_table, 'values')
        for name in names_categories:
            if categories[name][0][-1]==value[-1]:
                show_item_categories(box_categories, table_categories, categories,name)
                break
    except:pass
    refresh_Button_categories.config(text='Обновить данные', state='normal')
    refresh_Button_products.config(text='Обновить данные', state='normal')
    renew_all_button.config(text='Обновить данные', state='normal')
    add_statistics('category')
    if notify==1 and var1.get() == 1:
        message = 'Произошло обновление данных раздела "категории".'
        toast.show_toast(title='Обновление данных', msg=message, threaded=False, duration=5,icon_path='invisible.ico')
    if var3.get()==1 and prev_disc!=cur_disc:
        mark=''
        if prev_disc>cur_disc: mark='уменьшилось'
        else: mark='увеличилось'
        message = f'Количество товаров со скидкой '+mark+' на '+str(abs(prev_disc-cur_disc))+'.'
        toast.show_toast(title='Изменение скидок', msg=message, threaded=False, duration=5, icon_path='invisible.ico')
#запуск парсера раздела "товары"
def start_products(notify):
    refresh_Button_categories.config(text='Загрузка...',state='disabled')
    refresh_Button_products.config(text='Загрузка...',state='disabled')
    renew_all_button.config(text='Загрузка...',state='disabled')
    try:
        result = cur.execute("SELECT * FROM target_products;")
        names_products = list(element[0] for element in result.fetchall())
        products_values = []
        for product in names_products:
            if product.find('mvideo') >= 0: products_values.append(check_product_mvideo(product)[0])
            if product.find('dns') >= 0: products_values.append(check_product_dns(product)[0])
            if product.find('citilink') >= 0: products_values.append(check_product_citilink(product)[0])
        current_date = date.today().strftime('%d.%m.%Y')
        cur.execute(f"DELETE from products where date='{current_date}'")
        for element in products_values:
            element = element + [current_date]
            cur.execute("INSERT INTO products VALUES (?,?,?,?,?,?,?)", element)
        conn.commit()
        fill_products_variables()
    except: pass
    # заполнение таблицы products
    #################
    create_table_products()
    #############
    refresh_Button_categories.config(text='Обновить данные', state='normal')
    refresh_Button_products.config(text='Обновить данные', state='normal')
    renew_all_button.config(text='Обновить данные', state='normal')
    add_statistics('product')
    if notify == 1 and var1.get() == 1:
        message = 'Произошло обновление данных раздела "товары".'
        toast.show_toast(title='Обновление данных', msg=message, threaded=False, duration=5, icon_path='invisible.ico')
    products_changes()
#открытие таблицы с категориями, заполнение словаря categories
def fill_categories_variables():
    global categories,names_categories
    result = cur.execute("SELECT * FROM target_categories;")
    names_categories = list(element[0] for element in result.fetchall())
    categories={}
    for name in names_categories:
        try:
            result=cur.execute(f"SELECT * FROM categories where category='{name}';").fetchall()
            for i in range(len(result)):
                result[i]=result[i][1:-2]
            categories[name]=result
        except:
            pass
    sort_by_discount_categories()

#открытие таблицы с товарами, заполнение словаря products
def fill_products_variables():
    global names_products,names_by_link,products,dates_prices
    result = cur.execute("SELECT * FROM target_products;")
    names_products = list(element[0] for element in result.fetchall())
    names_by_link={}
    products={}
    dates_prices={}
    for name in names_products:
        try:
            result=cur.execute(f"SELECT * FROM products where link='{name}';").fetchall()
            products[name]=result[-1][1::]
            element=[]
            for i in result:
                element.append([i[2],i[-1]])
            dates_prices[name]=element
            names_by_link[name]=result[-1][3]
        except:pass

#сортировка товаров в разделе категорий по рамеру скидки
def sort_by_discount_categories():
    for name in names_categories:
        elements = categories[name]
        for elem in elements:
            elem=list(elem)
            elem[0]=int(elem[0])
        categories[name]=sorted(categories[name],reverse=True)
fill_categories_variables()
fill_products_variables()

#переключение между вкладками меню
def redirect(main,others,curr):
    for other in others:
        other.pack_forget()
    main.pack(fill="both", side="top", expand=True)
    for i in range(len(menu_buttons)):
        menu_buttons[i].configure(bg='white',fg='#373737')
    curr.configure(bg='#d2d2d2',fg='black')

#создание excel файла categories
def create_categories_excel():
    excel=Workbook()
    list1 = excel.create_sheet("Лист 1")
    del excel['Sheet']
    headers=['Скидка, %','Цена, ₽','Название','Магазин','Ссылка','Категория']
    list1.append(headers)
    list1.column_dimensions['A'].width = 10
    list1.column_dimensions['B'].width = 10
    list1.column_dimensions['C'].width = 60
    list1.column_dimensions['D'].width = 20
    list1.column_dimensions['E'].width = 60
    list1.column_dimensions['F'].width = 20
    for let in 'ABCDEF':
        list1[let+'1'].font = Font(bold=True)
    for name in names_categories:
        elements=categories[name]
        for elem in elements:
            list1.append(list(elem)+[name])
    excel.save('categories.xlsx')
    excel.close()
#создание excel файла products
def create_products_excel():
    excel=Workbook()
    list1 = excel.create_sheet("Лист 1")
    del excel['Sheet']
    headers=['Скидка, %','Цена, ₽','Название','Магазин','Ссылка','Дата']
    list1.append(headers)
    list1.column_dimensions['A'].width = 10
    list1.column_dimensions['B'].width = 10
    list1.column_dimensions['C'].width = 60
    list1.column_dimensions['D'].width = 20
    list1.column_dimensions['E'].width = 60
    list1.column_dimensions['F'].width = 15
    for let in 'ABCDEF':
        list1[let+'1'].font = Font(bold=True)
    elements=cur.execute("SELECT * FROM products;").fetchall()
    for element in elements:
        list1.append(element[1::])
    excel.save('products.xlsx')
    excel.close()

#создание окошка
window=Tk()
window.geometry("750x407")
window.title("")
window.resizable(False,False)
window.configure(bg="#000000")
window.iconbitmap('invisible.ico')
#создание меню
menu=Frame(window,width=600,height=40,bg='white')
menu.pack(fill=X)
discount_menu=Frame(window,width=600,height=450,bg='#d2d2d2')
tracked_menu=Frame(window,width=600,height=450,bg='#d2d2d2')
other_menu=Frame(window,width=600,height=450,bg='#d2d2d2')
discount_menu.pack(fill="both", side="top", expand=True)
all_frames=[discount_menu,tracked_menu,other_menu]
button_discount_menu=Button(menu,text='категории',bg="#d2d2d2",fg='black',activebackground='gray',command=lambda:redirect(discount_menu,all_frames,button_discount_menu),borderwidth=0,compound="left",font='MingLiU_HKSCS-ExtB 10')
button_discount_menu.pack(side=LEFT, expand=False)
button_tracked_menu=Button(menu,text='товары',bg="white",activebackground='gray',command=lambda:redirect(tracked_menu,all_frames,button_tracked_menu),borderwidth=0,compound="left",font='MingLiU_HKSCS-ExtB 10')
button_tracked_menu.pack(side=LEFT, expand=False)
button_other_menu=Button(menu,text='прочее',bg="white",activebackground='gray',command=lambda:redirect(other_menu,all_frames,button_other_menu),borderwidth=0,compound="left",font='MingLiU_HKSCS-ExtB 10')
button_other_menu.pack(side=LEFT, expand=False)
menu_buttons=[button_discount_menu,button_tracked_menu,button_other_menu]

''' discounts_menu - раздел "категории" '''
#добавить сохранённые товаров/категории из файла
def add_old(active,box):
    if len(active)>0:
        for elem in active:
            box.insert(END, elem)

#подсвечивание выбранных товаров/категорий
def color(id,box):
    elements=(box).get(0,END)
    for elem in range(len(elements)):
        box.itemconfig(elem,background='white')
    box.itemconfig(id, background='yellow')

#более читабельный вид для цены
def price(s):
    s=s[::-1]
    new_price=''
    for i in range(1,len(s)+1):
        if i%3!=0 or i==len(s):
            new_price+=s[i-1]
        else:
            new_price+=s[i-1]+'.'
    new_price='₽ '+new_price
    return new_price[::-1]

#добавить написанную категорию/продукт в список
def add_item(entry,active,type,box):
    cat = entry.get()
    if len(cat)>0 and cat not in active and len(active)<9:
        active.append(cat)
        if type=='category':
            cur.execute(f'INSERT INTO target_categories VALUES ("{cat}");')
        else:
            cur.execute(f'INSERT INTO target_products VALUES ("{cat}");')
        box.insert(END, cat)
        entry.delete(0, END)
    conn.commit()

#удалить категорию из списка
def del_item(box,active,table,dict,type):
    id = box.curselection()
    #удаление с таблицы
    if len(id)>0:
        current = box.get(id)
        active.remove(current)
        if box == box_products:
            show_del_item_products(0)
        else:
            try:
                if dict[current][0][2] == table.item(table.get_children()[0]).get('values')[2]:
                    for row in table.get_children():
                        table.delete(row)
            except:
                pass
        #удаление с базы данных
        if type=='categories':
            cur.execute(f'DELETE FROM target_categories WHERE category="{current}";')
            cur.execute(f'DELETE FROM categories WHERE category="{current}";')
        else:
            cur.execute(f'DELETE FROM target_products WHERE product="{current}";')
            cur.execute(f'DELETE FROM products WHERE link="{current}";')
        conn.commit()
        select = list(box.curselection())
        select.reverse()
        for i in select:
            box.delete(i)

#выбрать отоброжаемую категорию в таблице
def show_item_categories(box,table,dict,cur_category):
    if cur_category==0:
        id=box.curselection()
        color(id,box)
        cur_category=box.get(id)
    try:
        for row in table.get_children():
            table.delete(row)
        for row in (dict[cur_category]):
            row=list(row)
            row[0]=str(row[0])
            row[0] = row[0].replace(' %','') + ' %'
            row[1]=str(row[1])
            row[1]=row[1].replace('.','').replace(' ₽', '')
            row[1] = price(str(row[1]))
            table.insert('', END, values=row)
    except:
        pass

#открыть сылку на товар в бразуре
def open_url(table):
    id=table.focus()
    if table==table_categories:
        url=table.item(id,'values')[-1]
    else:
        url=table.item(id,'values')[-2]
    webbrowser.open(url,new=2)

#открыть таблицу в эксель
def open_excel(type):
    if type=='categories': create_categories_excel()
    else: create_products_excel()
    os.startfile(type+'.xlsx')

#разделение поля на части
f1_categories = Frame(discount_menu,bg='#d2d2d2')
f1_categories.pack(side=LEFT,fill=Y)
f2_categories=Frame(discount_menu,bg='#d2d2d2')
f2_categories.pack(side=LEFT,fill=BOTH)
f2_1_categories=Frame(f2_categories,bg='#d2d2d2')
f2_1_categories.pack(pady=12)
f2_2_categories=Frame(f2_categories,bg='#d2d2d2')
f2_2_categories.pack()

#загрузка сохраннённых категорий из файла в список категорий
box_categories = Listbox(f1_categories,selectmode=SINGLE,relief="flat")
add_old(names_categories,box_categories)
box_categories.pack(pady=12,padx=10)
#выделение цвета первого элемнта
try:color(0,box_categories)
except:pass

#заполнение таблицы discount_menu
heads=['Скидка','Цена','Название','Магазин']
table_categories=ttk.Treeview(f2_1_categories,show='headings',selectmode="browse",height=15)
table_categories['columns']=heads
for header in heads:
    table_categories.heading(header,text=header,anchor='center')
    table_categories.column(header,anchor='center')
try:
    for row in (categories[names_categories[0]]):
        row=list(row)
        row[0] = str(row[0])
        row[0] = row[0] + ' %'
        row[1] = price(str(row[1]))
        table_categories.insert('', END, values=row)
except:pass
table_categories.column("Скидка", minwidth=0, width=50, stretch=NO)
table_categories.column("Цена", minwidth=0, width=50, stretch=NO)
table_categories.column("Название", minwidth=0, width=400, stretch=NO)
table_categories.column("Магазин", minwidth=0, width=75, stretch=NO)
table_categories.pack(expand=YES,fill=BOTH,side=LEFT)
#ползунок для таблицы
scroll_pane_categ=ttk.Scrollbar(f2_1_categories,command=table_categories.yview)
table_categories.configure(yscrollcommand=scroll_pane_categ.set)
scroll_pane_categ.pack(side=LEFT,fill=Y)

#кнопки для таблицы
open_Button=Button(f2_2_categories,text="Открыть страницу товара", command=lambda:open_url(table_categories))
open_Button.pack(side=LEFT,padx=3)
excel_Button=Button(f2_2_categories,text="Открыть таблицу в Excel", command=lambda:open_excel('categories'))
excel_Button.pack(side=LEFT,padx=3)
refresh_Button_categories=Button(f2_2_categories,text='Обновить данные', command=lambda:threading.Thread(target=start_categories,args=(1,)).start())
refresh_Button_categories.pack(side=LEFT,padx=3)
#кнопки для списка
entry_categories = Entry(f1_categories)
entry_categories.pack(anchor=N,pady=3)
add_Button_categories=Button(f1_categories, text="Добавить", command=lambda:add_item(entry_categories,names_categories,'category',box_categories),width=16)
add_Button_categories.pack(pady=1,anchor=CENTER)
del_Button_categories=Button(f1_categories, text="Удалить", command=lambda:del_item(box_categories,names_categories,table_categories,categories,'categories'),width=16)
del_Button_categories.pack(pady=1,anchor=CENTER)
show_Button_categories=Button(f1_categories, text="Показать", command=lambda:show_item_categories(box_categories,table_categories,categories,0),width=16)
show_Button_categories.pack(pady=1,anchor=CENTER)



''' tracked_menu - раздел "товары" '''

#открытие графика изменения цены
def create_graph(table):
    fig = pylab.gcf()
    fig.canvas.manager.set_window_title('График изменения цены')
    id = table.focus()
    item = table.item(id, 'values')
    item=item[4]
    mass=dates_prices[item]
    xs=[]
    ys=[]
    for i in range(len(mass)):
        d=mass[i][-1]
        xs.append(d[:d.rfind('.')])
        ys.append(mass[i][0])
    ax=fig.add_subplot()
    ax.plot()
    if len(mass)>=10:
        max_x_in_row=10
    else:
        if len(mass)==1:max_x_in_row=0
        else:max_x_in_row=len(mass)
    ax.xaxis.set_major_locator(LinearLocator(max_x_in_row))
    ax.yaxis.set_major_locator(LinearLocator(7))
    plt.xlabel('Дата')
    plt.ylabel('Цена')
    name=names_by_link[item]
    if len(name)>47:name=name[0:47]+'...'
    plt.title(name)
    plt.plot(xs,ys)
    plt.show()

#выделение/удаление товары из выбранной ссылки
def show_del_item_products(mark):
    ids_table = table_products.get_children()
    id_box = box_products.curselection()
    if mark == 1:
        color(id_box, box_products)
        #вставка значенийj
    if len(id_box) > 0:
        link = box_products.get(id_box)
        for id_table in ids_table:
            table_products.tag_configure(id_table,background='white')
            if table_products.item(id_table, 'values')[-2]==link:
                if mark==1:
                    table_products.tag_configure(id_table,background='yellow')
                else:
                    table_products.delete(id_table)
def create_table_products():
    global table_products
    table_products.destroy()
    # заполнение таблицы products_menu
    heads = ['Скидка', 'Цена', 'Название']
    table_products = ttk.Treeview(f_1_2_products, show='headings', selectmode="browse", height=10)
    style = ttk.Style()
    style.configure("Treeview", background='white')
    table_products['columns'] = heads
    for header in heads:
        table_products.heading(header, text=header, anchor='center')
        table_products.column(header, anchor='center')
    count = 0
    # заполнение таблицы products
    try:
        for product in names_products:
            count += 1
            id_tag = str(count)
            while len(id_tag) < 3:
                id_tag = '0' + id_tag
            row = products[product]
            row = list(row)
            row[0] = str(row[0])
            row[0] = row[0] + ' %'
            row[1] = price(str(row[1]))
            table_products.insert('', END, values=row, tags=('I' + id_tag))
    except:
        pass
    table_products.column("Скидка", minwidth=0, width=50, stretch=NO)
    table_products.column("Цена", minwidth=0, width=75, stretch=NO)
    table_products.column("Название", minwidth=0, width=454, stretch=NO)
    table_products.pack(expand=YES, fill=BOTH, side=LEFT, pady=5, padx=5)
    elements = (box_products).get(0, END)
    for elem in range(len(elements)):
        box_products.itemconfig(elem, background='white')
#разделения поля на части
f1_products = Frame(tracked_menu,bg='#d2d2d2')
f1_products.pack(side=LEFT,fill=Y)
f2_products=Frame(tracked_menu,bg='#d2d2d2')
f2_products.pack(side=LEFT,fill=Y)
f1_1_products=Frame(f1_products,bg='#d2d2d2')
f1_1_products.pack()
f_1_2_products=Frame(f1_products,bg='#d2d2d2')
f_1_2_products.pack()
f2_1_products=Frame(f2_products,bg='#d2d2d2')
f2_1_products.pack(pady=20)
f2_2_products=Frame(f2_products,bg='#d2d2d2')
f2_2_products.pack(pady=37)
#загрузка сохраннённых категорий из файла в список категорий
box_products = Listbox(f1_1_products,selectmode=SINGLE,relief="flat",width=96,height=9)
add_old(names_products,box_products)
box_products.pack(pady=7,padx=7)
table_products = ttk.Treeview(f_1_2_products, show='headings', selectmode="browse", height=10)
create_table_products()
#кнопки для таблицы
graph_button=Button(f2_2_products,text='График изменения цены', command=lambda:create_graph(table_products),width=20)
graph_button.pack(pady=4,padx=5)
open_Button=Button(f2_2_products,text="Открыть страницу товара", command=lambda:open_url(table_products),width=20)
open_Button.pack(pady=4)
excel_Button=Button(f2_2_products,text="Открыть таблицу в Excel", command=lambda:open_excel('products'),width=20)
excel_Button.pack(pady=4)
refresh_Button_products=Button(f2_2_products,text='Обновить данные', command=lambda:threading.Thread(target=start_products,args=(1,)).start())
refresh_Button_products.pack(pady=4)
#сохдание кнопок и поля для списка категорий
entry_products = Entry(f2_1_products)
entry_products.pack(anchor=N,pady=5,fill=X)
add_Button_products=Button(f2_1_products, text="Добавить", command=lambda:add_item(entry_products,names_products,'product',box_products),width=16)
add_Button_products.pack(anchor=CENTER,pady=2)
del_Button_products=Button(f2_1_products, text="Удалить", command=lambda:del_item(box_products,names_products,table_products,products,'products'),width=16)
del_Button_products.pack(anchor=CENTER,pady=2)
show_Button_products=Button(f2_1_products, text="Показать", command=lambda:show_del_item_products(1),width=16)
show_Button_products.pack(anchor=CENTER,pady=2)

''' other_menu - раздел "прочее" '''
#сохраниение периодичности и её запуск
def timer():
    global job
    global flag
    flag=0
    time.sleep(1.1)
    mark=combo_timer.get()
    g=entry_timer.get()
    if g=='':count=0
    else: count=int(g)
    sec=0
    if count>0:
        if mark=='минут':
            sec=count*60
        if mark=='часов':
            sec=count*60*60
        if mark == 'дней':
            sec=count*60*60*24
        flag=1
        threading.Thread(target=start_period,args=(sec,)).start()
    file = open('save.txt', 'r')
    status = file.readlines()
    file.close()
    for i in range(4):
        status[i] = status[i].split(' ')
    file = open('save.txt', 'w')
    s = ''
    for i in range(4):
        if i != 3:
            s += status[i][0] + ' ' + status[i][1]
        else:
            s+=combo_values[mark]+' '+str(count)
    file.write(s)
    file.close()

#скачать базу данных
def click_link(event):
    try:
        path=filedialog.asksaveasfilename(filetypes=[('Data Base File (*.db)','*.db')],defaultextension='.db',title='Сохранение файла',initialfile='DataBase.db')
        shutil.copyfile('discounts.db',path)
    except:pass

#удаление базы данных
def delete_data_base():
    choice=messagebox.askyesno('Удаление','Удалить всю информацию о товарах?')
    if choice==True:
        tables_database=['categories','products','statistics','target_categories','target_products']
        for table in tables_database: cur.execute(f"DELETE from {table};")
        conn.commit()
        tables_sections=[table_categories,table_products,table_statistics,table_all_statistics]
        for table in tables_sections:
            for row in table.get_children():
                table.delete(row)
        table_all_statistics.insert('', END, values=['Товары', '', '', ''])
        table_all_statistics.insert('', END, values=['Категории', '', '', ''])
        boxes_sections=[box_categories,box_products]
        try:
            for box in boxes_sections:
                elements=box.get(0, END)
                for element in elements:
                    box.delete(0)
        except:pass

#удаление статистики
def delete_statistics():
    for element in table_all_statistics.get_children():
        table_all_statistics.delete(element)
    for element in table_statistics.get_children():
        table_statistics.delete(element)
    cur.execute("DELETE from statistics;")
    conn.commit()

#заполенние таблицы полной статистики
def refresh_all_statistics():
    shops=['dns-shop.ru','mvideo.ru','citilink.ru']
    sections=['Товары','Категории']
    rows=[]
    for section in sections:
        row=[section]
        for shop in shops:
            s=0
            element=cur.execute('SELECT value FROM statistics where (shop=? and section= ?)',[shop,section]).fetchall()
            for i in range(len(element)):
                s+=element[i][0]
            row.append(s)
        rows.append(row)
    for element in table_all_statistics.get_children():
        table_all_statistics.delete(element)
    try:
        for row in rows:
            table_all_statistics.insert('', END, values=row)
    except:pass

#заполенние таблицы статистики
def refresh_statistics():
    rows = cur.execute('SELECT * FROM statistics').fetchall()
    for i in range(len(rows)): rows[i] = list(rows[i])
    for row in table_statistics.get_children():
        table_statistics.delete(row)
    try:
        for row in rows:
            table_statistics.insert('', END, values=row)
    except:
        pass
    refresh_all_statistics()

#нахождение информации о недавных процессах парсинга
def add_statistics(type):
    result = cur.execute('SELECT date FROM categories').fetchall()
    for i in range(len(result)): result[i] = result[i][0]
    dates = list(set(result))
    result = cur.execute('SELECT shop FROM shops').fetchall()
    for i in range(len(result)): result[i] = result[i][0]
    shops = result
    for date in dates:
        for shop in shops:
            if type=='category':
                result = cur.execute('SELECT * FROM categories where (date=? and shop=?);', (date, shop)).fetchall()
                values = (date, 'Категории', shop, len(result))
                cur.execute('INSERT INTO statistics VALUES (?,?,?,?);', values)
            else:
                result = cur.execute('SELECT * FROM products where (date=? and shop=?);', (date, shop)).fetchall()
                values = (date, 'Товары', shop, len(result))
                cur.execute('INSERT INTO statistics VALUES (?,?,?,?);', values)
    conn.commit()
    refresh_statistics()

#сохранить флажки уведомлений
def save_checkbutton():
    b1=str(var1.get())
    b2=str(var2.get())
    b3=str(var3.get())
    b=[b1,b2,b3]
    file=open('save.txt','r',encoding='UTF-8')
    status=file.readlines()
    file.close()
    for i in range(4):
        status[i]=status[i].split(' ')
    file=open('save.txt','w')
    s=''
    for i in range(4):
        if i!=3:
            s+=status[i][0]+' '+b[i]+'\n'
        else:
            s+=status[i][0]+' '+status[i][1]
    file.write(s)
    file.close()
def products_changes():
    for name in names_products:
        element=dates_prices[name]
        if len(element)>1 and element[-1][0]!=element[-2][0]:
            if var2.get() == 1:
                message = 'Цена на некоторые товары изменилась.'
                toast.show_toast(title='Изменение цены', msg=message, threaded=False, duration=5, icon_path='invisible.ico')
def start_period(sec):
    while True:
        for i in range(sec):
            time.sleep(1)
            if flag==0:
                return 0
        start_all()
def off_the_preiod():
    global flag
    flag=0
    entry_timer.delete('0', END)
    entry_timer.insert(END, '0')
    timer()
    entry_timer.delete('0',END)

#разделение поля
f1_other = Frame(other_menu,bg='#d2d2d2')
f1_other.pack(side=LEFT,fill=Y)
f2_other=Frame(other_menu,bg='#d2d2d2')
f2_other.pack(side=LEFT,fill=Y)
f1_1_other=LabelFrame(f1_other,bg='#f0f0f0',text='Настройка уведомлений',borderwidth=1,relief='solid',font=(10))
f1_1_other.pack(padx=5,pady=5,fill=X)
f1_2_other=LabelFrame(f1_other,bg='#f0f0f0',text='Настройка сбора информации',borderwidth=1,relief='solid',font=(10))
f1_2_other.pack(padx=5,pady=5,fill=X)
f1_3_other=LabelFrame(f1_other,bg='#f0f0f0',text='Настройка базы данных',borderwidth=1,relief='solid',font=(10),width=9999)
f1_3_other.pack(padx=5,pady=5,fill=X)

#настройка уведмолений
var1=IntVar()
var2=IntVar()
var3=IntVar()
check1=Checkbutton(f1_1_other,text='Получать уведомления по окончании сбора информации',font='TkDefaultFont 10',variable=var1)
check2=Checkbutton(f1_1_other,text='Получать уведомления при изменении цены на товар',font='TkDefaultFont 10',variable=var2)
check3=Checkbutton(f1_1_other,text='Получать уведомления при получении новых скидок',font='TkDefaultFont 10',variable=var3)
check1.pack(anchor=W,pady=1)
check2.pack(anchor=W,pady=1)
check3.pack(anchor=W,pady=1)
#проставление сохраненных галочек checkkbutton
file=open('save.txt','r')
status=file.readlines()
file.close()
checkbuttons=[check1,check2,check3]
for i in range(3):
    status[i]=status[i].split(' ')
    if status[i][1]=='1\n': checkbuttons[i].select()
save_button_notify=Button(f1_1_other,text='Сохранить изменения',font='TkDefaultFont 10',command=lambda:save_checkbutton())
save_button_notify.pack(pady=2)

#Настройка сбора информации
timer_frame=Frame(f1_2_other)
timer_frame.pack(pady=3,anchor=W)
label_timer=Label(timer_frame,text='Периодичность сбора данных',font='TkDefaultFont 10')
label_timer.pack(side=LEFT,pady=4,padx=5)
entry_timer=Entry(timer_frame,width=10)
entry_timer.pack(side=LEFT,padx=2)
combo_timer=ttk.Combobox(timer_frame,values=('минут','часов','дней'),state="readonly",width=7,font='TkDefaultFont 10')
combo_values={'минут':'0','часов':'1','дней':'2'}
combo_timer.pack(side=LEFT,padx=2)
save_button_timer=Button(timer_frame,text='Сохранить', command=lambda:timer(),font='TkDefaultFont 10')
save_button_timer.pack(side=LEFT,padx=2)
#установка значений
file=open('save.txt')
status=file.readlines()[-1].split()
file.close()
secti1=int(status[-1])
secti2=(status[0])
#установка значения в combobox
combo_timer.current(int(secti2))
#установка значения в поле и запуск периодического парсинга
flag=0
if secti1>0:
    mark = combo_timer.get()
    sec=0
    if mark == 'минут':
        sec = secti1 * 60
    if mark == 'часов':
        sec = secti1 * 60 * 60
    if mark == 'дней':
        sec = secti1 * 60 * 60 * 24
    entry_timer.insert(END, str(secti1))
    if sec>0:
        flag = 1
        threading.Thread(target=start_period, args=(sec,)).start()
off_the_preiod_Frame=Frame(f1_2_other)
off_the_preiod_Frame.pack(pady=3,anchor=W)
renew_Frame=Frame(f1_2_other)
renew_Frame.pack(pady=3,anchor=W)
off_the_preiod_label=Label(off_the_preiod_Frame,text="Отключить периодический сбор данных",font='TkDefaultFont 10')
off_the_preiod_label.pack(side=LEFT,padx=5)
off_the_preiod_button=Button(off_the_preiod_Frame,text="Отключить", command=lambda:off_the_preiod(),font='TkDefaultFont 10')
off_the_preiod_button.pack(side=LEFT)
renew_all_label=Label(renew_Frame,text="Собрать информацию со всех разделов",font='TkDefaultFont 10')
renew_all_label.pack(side=LEFT,padx=5)
renew_all_button=Button(renew_Frame,text="Обновить данные", command=lambda:threading.Thread(target=start_all).start(),font='TkDefaultFont 10')
renew_all_button.pack(side=LEFT)

#настройка базы данных
f1_3_other_1=Frame(f1_3_other)
f1_3_other_1.pack(pady=3,anchor=W)
f1_3_other_2=Frame(f1_3_other)
f1_3_other_2.pack(pady=3,anchor=W)
poin_on_link_label=Label(f1_3_other_2,text='Скачать базу данных ',font='TkDefaultFont 10')
poin_on_link_label.pack(side=LEFT,padx=5)
link_download = Label(f1_3_other_2,text='Скачать файл', foreground="#0000ff")
f = font.Font(link_download, link_download.cget("font"))
f.configure(underline=True)
link_download.configure(font=f)
link_download.bind('<Button-1>', click_link)
link_download.pack(side=LEFT,padx=2)
poin_on_delete_button=Label(f1_3_other_1,text='Удалить информацию с базы данных ',font='TkDefaultFont 10')
poin_on_delete_button.pack(side=LEFT,padx=5)
delete_database_button=Button(f1_3_other_1,text='Удалить данные', command=lambda:delete_data_base(),font='TkDefaultFont 10')
delete_database_button.pack(side=LEFT,padx=2)

#заполнение таблицы операций
f2_1_other=Frame(f2_other,bg="#d2d2d2")
f2_1_other.pack()
f2_2_other=Frame(f2_other,bg="#d2d2d2")
f2_2_other.pack(fill=X)
heads=['Дата','Раздел','Магазин','Количество']
table_statistics=ttk.Treeview(f2_1_other,show='headings',selectmode="browse",height=12,)
style=ttk.Style()
style.configure("Treeview",background='white')
table_statistics['columns']=heads
for header in heads:
    table_statistics.heading(header,text=header,anchor='center')
    table_statistics.column(header,anchor='center')
table_statistics.column("Дата", minwidth=0, width=75, stretch=NO)
table_statistics.column("Раздел", minwidth=0, width=75, stretch=NO)
table_statistics.column("Магазин", minwidth=0, width=75, stretch=NO)
table_statistics.column("Количество", minwidth=0, width=75, stretch=NO)
table_statistics.pack(expand=YES,fill=BOTH,side=LEFT,pady=5)
#ползунок для таблицы
scroll_pane_stat=ttk.Scrollbar(f2_1_other,command=table_statistics.yview)
table_statistics.configure(yscrollcommand=scroll_pane_stat.set)
scroll_pane_stat.pack(side=LEFT,fill=Y,pady=5)

#таблица полной статистики
heads=['','dns-shop.ru','mvideo.ru','citilink.ru']
table_all_statistics=ttk.Treeview(f2_2_other,show='headings',selectmode="browse",height=2)
style=ttk.Style()
style.configure("Treeview",background='white')
table_all_statistics['columns']=heads
for header in heads:
    table_all_statistics.heading(header,text=header,anchor='center')
    table_all_statistics.column(header,anchor='center')
table_all_statistics.column("", minwidth=0, width=75, stretch=NO)
table_all_statistics.column("mvideo.ru", minwidth=0, width=75, stretch=NO)
table_all_statistics.column("dns-shop.ru", minwidth=0, width=75, stretch=NO)
table_all_statistics.column("citilink.ru", minwidth=0, width=75, stretch=NO)
table_all_statistics.pack(expand=YES,side=LEFT,pady=5,anchor=W)
table_all_statistics.insert('', END, values=['Товары','','',''])
table_all_statistics.insert('', END, values=['Категории','','',''])
refresh_statistics()

#кнопка для очистки статистики
delete_statistics_button=Button(f2_other,text='Очистить статистику', command=lambda:delete_statistics(),font='TkDefaultFont 10')
delete_statistics_button.pack()
window.protocol("WM_DELETE_WINDOW", window_closed)
window.mainloop()
conn.close()